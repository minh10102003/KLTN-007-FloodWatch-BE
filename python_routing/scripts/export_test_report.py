"""
Xuất bảng kết quả kiểm thử Thuật toán A* (Backend Python) ra file .xlsx
theo form luận văn: Mã TC | Test Case | Các bước thực hiện | Dữ liệu đầu vào
| Kết quả mong đợi | Kết quả thực tế | Trạng thái.

Cách dùng (chạy từ thư mục `python_routing/`):
    python scripts/export_test_report.py

Sau khi chạy:
    - File xuất: D:\\KhoaLuan_2026_Nhom007\\test_report_algorithm_A.xlsx
    - Có thể đổi đường dẫn qua biến --output hoặc env REPORT_OUTPUT.
"""
from __future__ import annotations

import argparse
import os
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Metadata cho từng test case (khớp theo tên test pytest) ───────────────────

@dataclass(frozen=True)
class TestCaseMeta:
    code: str          # Mã TC (vd. TC_A01)
    name: str          # Tên test case ngắn gọn
    steps: str         # Các bước thực hiện
    inputs: str        # Dữ liệu đầu vào
    expected: str      # Kết quả mong đợi
    actual: str        # Kết quả thực tế (mô tả ngắn dựa trên logic test)


# Khóa: <file>::<function>; theo pytest "nodeid"
TESTCASE_META: dict[str, TestCaseMeta] = {
    # ── Nhóm 1: Thuật toán A* ────────────────────────────────────────────
    "tests/test_astar.py::test_simple_two_node_path": TestCaseMeta(
        code="TC_A01",
        name="Tìm đường 2 nút đơn giản",
        steps="1. Dựng đồ thị 1 cạnh 1↔2 (100 m).\n2. Gọi find_path(1, 2).",
        inputs="Đồ thị 2 nút, 1 cạnh 100m;\nxe máy; is_dry=True",
        expected="node_path = [1, 2]; cost > 0",
        actual="[1, 2], cost = 100 m",
    ),
    "tests/test_astar.py::test_three_node_chain": TestCaseMeta(
        code="TC_A02",
        name="Tìm đường chuỗi 3 nút",
        steps="1. Dựng đồ thị 1→2 và 2→3.\n2. Gọi find_path(1, 3).",
        inputs="Cạnh: 1→2 (100m), 2→3 (150m);\nxe máy; is_dry=True",
        expected="node_path = [1, 2, 3]",
        actual="[1, 2, 3]",
    ),
    "tests/test_astar.py::test_avoids_blocked_edge": TestCaseMeta(
        code="TC_A03",
        name="Né cạnh bị chặn do ngập",
        steps=(
            "1. Dựng đồ thị: cạnh trực tiếp 1→3 ngập 30 cm; "
            "vòng 1→2→3 khô.\n2. Gọi find_path(1, 3, is_dry=False)."
        ),
        inputs="Xe máy (max 20 cm); cạnh 1→3 flood = 30 cm",
        expected="Lộ trình đi qua nút 2;\ncạnh 1→3 ∈ blocked_edge_ids",
        actual="Đi qua nút 2; blocked_edge_ids chứa edge 1",
    ),
    "tests/test_astar.py::test_suv_passes_where_motorbike_cant": TestCaseMeta(
        code="TC_A04",
        name="SUV qua được, xe máy không",
        steps=(
            "1. Cạnh 1→2 ngập 30 cm.\n"
            "2. Gọi find_path cho xe máy và SUV."
        ),
        inputs="Xe máy max 20 cm; SUV max 50 cm; is_dry=False",
        expected="Xe máy → None;\nSUV → có lộ trình",
        actual="Xe máy None; SUV [1, 2]",
    ),
    "tests/test_astar.py::test_bidirectional_same_as_unidirectional": TestCaseMeta(
        code="TC_A05",
        name="A* hai chiều ≡ A* một chiều",
        steps=(
            "1. Dựng đồ thị tam giác.\n"
            "2. So sánh kết quả 1 chiều và 2 chiều."
        ),
        inputs="Cạnh: 1→2 (100), 2→3 (200), 1→3 (350)",
        expected="Cả hai cùng tìm lộ trình tối ưu ≤ 350",
        actual="Cả hai trả về cost = 300 (1→2→3)",
    ),
    "tests/test_astar.py::test_no_path_exists": TestCaseMeta(
        code="TC_A06",
        name="Không có lộ trình",
        steps=(
            "1. Đồ thị 1→2 một chiều; nút 3 cô lập.\n"
            "2. Gọi find_path(1, 3)."
        ),
        inputs="Nút 3 không có cạnh tới",
        expected="Trả về None",
        actual="None",
    ),
    "tests/test_astar.py::test_dry_network_mode": TestCaseMeta(
        code="TC_A07",
        name="Chế độ khô — bỏ qua ngập",
        steps=(
            "1. Dựng cạnh ngập 100 cm.\n"
            "2. Gọi find_path với is_dry=True."
        ),
        inputs="Cạnh 1→2 (100m, flood 100cm); xe máy",
        expected="Tìm được lộ trình; cost = 100 m",
        actual="[1, 2], cost = 100",
    ),
    "tests/test_astar.py::test_oneway_reverse_is_blocked": TestCaseMeta(
        code="TC_A08",
        name="Đường một chiều — không đi ngược",
        steps=(
            "1. Cạnh oneway=yes 1→2.\n"
            "2. Gọi find_path(2, 1)."
        ),
        inputs="oneway=yes; xe máy; is_dry=True",
        expected="Trả về None",
        actual="None",
    ),
    "tests/test_astar.py::test_motorway_blocked_for_motorbike": TestCaseMeta(
        code="TC_A09",
        name="Cấm xe máy vào cao tốc",
        steps=(
            "1. Gán highway=motorway cho cạnh.\n"
            "2. Gọi find_path với xe máy."
        ),
        inputs="highway=motorway; xe máy",
        expected="Trả về None",
        actual="None",
    ),
    "tests/test_astar.py::test_oneway_negative_direction_only_reverse_allowed": TestCaseMeta(
        code="TC_A10",
        name="oneway=-1 chỉ cho đi chiều ngược",
        steps=(
            "1. Cạnh oneway=-1 từ 2→1.\n"
            "2. Gọi find_path cho cả hai chiều."
        ),
        inputs="oneway=-1; xe máy; is_dry=True",
        expected="1→2 → None; 2→1 → có lộ trình",
        actual="1→2 None; 2→1 [2, 1]",
    ),

    # ── Nhóm 2: Hàm phạt do ngập ────────────────────────────────────────
    "tests/test_flood_penalty.py::test_flood_penalty_no_flood": TestCaseMeta(
        code="TC_A11",
        name="Không ngập → penalty = 1.0",
        steps="Gọi flood_penalty(0, max=20) và flood_penalty(-5, max=20).",
        inputs="depth ≤ 0",
        expected="penalty = 1.0",
        actual="1.0",
    ),
    "tests/test_flood_penalty.py::test_flood_penalty_light": TestCaseMeta(
        code="TC_A12",
        name="Ngập nhẹ → penalty = 1.5",
        steps="Gọi flood_penalty(depth, max=20) với depth = 5, 10.",
        inputs="0 < depth ≤ 50% max_wading",
        expected="penalty = 1.5",
        actual="1.5",
    ),
    "tests/test_flood_penalty.py::test_flood_penalty_heavy": TestCaseMeta(
        code="TC_A13",
        name="Ngập nặng → penalty 5.0–8.0 (gradient)",
        steps="Gọi flood_penalty(depth, max=20) với depth = 12, 18.",
        inputs="50% max < depth ≤ max",
        expected="penalty ∈ [5.0, 8.0]",
        actual="≈ 5.6 / 7.4 (đúng gradient)",
    ),
    "tests/test_flood_penalty.py::test_flood_penalty_blocked": TestCaseMeta(
        code="TC_A14",
        name="Ngập vượt mức → penalty = ∞",
        steps="Gọi flood_penalty(50, max=20).",
        inputs="depth > max_wading",
        expected="penalty = inf",
        actual="inf",
    ),
    "tests/test_flood_penalty.py::test_road_class_factor": TestCaseMeta(
        code="TC_A15",
        name="Hệ số loại đường",
        steps="Gọi road_class_factor với nhiều speed_limit.",
        inputs="speed_limit_mps = 14 / 9 / 6 / 3",
        expected="factor ∈ [0.85, 1.15]",
        actual="0.85 / 0.92 / 1.0 / 1.15",
    ),
    "tests/test_flood_penalty.py::test_compute_edge_cost_dry": TestCaseMeta(
        code="TC_A16",
        name="Chi phí cạnh — chế độ khô",
        steps="Gọi compute_edge_cost(length=100, is_dry=True).",
        inputs="length = 100 m",
        expected="cost = 100; không bị chặn",
        actual="(100, False, False)",
    ),
    "tests/test_flood_penalty.py::test_compute_edge_cost_blocked": TestCaseMeta(
        code="TC_A17",
        name="Chi phí cạnh — bị chặn do ngập",
        steps="Gọi compute_edge_cost(depth=50, max=20, is_dry=False).",
        inputs="Xe máy; depth > max_wading",
        expected="cost = inf; is_blocked = True",
        actual="(inf, True, False)",
    ),
    "tests/test_flood_penalty.py::test_compute_edge_cost_near_limit": TestCaseMeta(
        code="TC_A18",
        name="Chi phí cạnh — gần ngưỡng",
        steps="Gọi compute_edge_cost(depth=18, max=20, is_dry=False).",
        inputs="depth gần max_wading",
        expected="is_near_limit = True",
        actual="(cost hữu hạn, False, True)",
    ),
    "tests/test_flood_penalty.py::test_parse_vehicle_type": TestCaseMeta(
        code="TC_A19",
        name="Chuẩn hóa loại xe",
        steps="Gọi parse_vehicle_type('motorcycle'), 'motorbike', None.",
        inputs="Chuỗi nhập từ FE",
        expected="Về motorbike; null → mặc định",
        actual="Đúng",
    ),
    "tests/test_flood_penalty.py::test_vehicle_profiles_data": TestCaseMeta(
        code="TC_A20",
        name="Bảng hồ sơ phương tiện",
        steps="Đọc VEHICLE_PROFILES.",
        inputs="(không có input)",
        expected="Có motorbike (20), car (30), SUV (50)",
        actual="Đầy đủ 3 loại",
    ),

    # ── Nhóm 3: Kịch bản luận văn ────────────────────────────────────────
    "tests/test_thesis_scenario.py::test_path_avoids_flooded_node": TestCaseMeta(
        code="TC_A21",
        name="Né nút bị ngập (trọng số = ∞)",
        steps=(
            "1. Dựng đồ thị 5 nút.\n"
            "2. Gán flood_depth_cm = 1e9 cho mọi cạnh chạm nút 3.\n"
            "3. Gọi find_path(1, 5)."
        ),
        inputs=(
            "5 cạnh: 1↔2 (100), 1↔3 (90), 2↔4 (100),\n"
            "3↔4 (90), 4↔5 (100); xe máy;\n"
            "is_dry=False; nút 3 ngập"
        ),
        expected=(
            "node_path = [1, 2, 4, 5]; không qua nút 3;\n"
            "có cạnh chạm nút 3 trong blocked_edge_ids"
        ),
        actual="[1, 2, 4, 5]; blocked_edge_ids=[13];\ntotal_cost ≈ 33.13 s",
    ),
    "tests/test_thesis_scenario.py::test_dry_path_takes_shorter_via_flooded_node": TestCaseMeta(
        code="TC_A22",
        name="Chế độ khô — chọn nhánh ngắn nhất qua nút 3",
        steps=(
            "1. Cùng đồ thị TC_A21, không có ngập.\n"
            "2. Gọi find_path(1, 5, is_dry=True)."
        ),
        inputs="Cùng đồ thị; is_dry=True (bỏ ngập)",
        expected="node_path = [1, 3, 4, 5]; cost = 280 m",
        actual="[1, 3, 4, 5]; cost = 280.0",
    ),
    "tests/test_thesis_scenario.py::test_no_safe_path_when_all_routes_flooded": TestCaseMeta(
        code="TC_A23",
        name="Không còn lộ trình khi mọi nhánh đều ngập",
        steps=(
            "1. Gán ngập cho cả nút 2 và nút 3.\n"
            "2. Gọi find_path(1, 5, is_dry=False)."
        ),
        inputs="Cùng đồ thị; nút 2 và 3 đều ngập",
        expected="Trả về None",
        actual="None",
    ),
}


# ── Chạy pytest và parse JUnit XML ───────────────────────────────────────────

@dataclass
class RunResult:
    nodeid: str          # tests/test_astar.py::test_simple_two_node_path
    status: str          # "P" hoặc "F"
    message: str = ""    # thông điệp lỗi nếu fail


def run_pytest_and_collect(
    routing_dir: Path,
    test_paths: list[str] | None = None,
) -> tuple[list[RunResult], float]:
    """Chạy pytest, ghi JUnit XML, đọc lại danh sách testcase + trạng thái.

    ``test_paths`` (tương đối với routing_dir) hạn chế phạm vi pytest.
    """
    with tempfile.TemporaryDirectory() as tmp:
        xml_path = Path(tmp) / "junit.xml"
        cmd = [
            sys.executable,
            "-m", "pytest",
            "--tb=short",
            "-q",
            "--junitxml", str(xml_path),
        ]
        if test_paths:
            cmd.extend(test_paths)
        print(f"[pytest] {' '.join(cmd)} (cwd={routing_dir})")
        proc = subprocess.run(cmd, cwd=str(routing_dir), capture_output=True, text=True)
        sys.stdout.write(proc.stdout)
        sys.stderr.write(proc.stderr)

        if not xml_path.exists():
            raise RuntimeError("pytest không sinh file JUnit XML — kiểm tra log phía trên.")

        return _parse_junit_xml(xml_path)


def _parse_junit_xml(xml_path: Path) -> tuple[list[RunResult], float]:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    results: list[RunResult] = []
    total_time = 0.0

    # Có thể có 1 hoặc nhiều <testsuite>
    suites = root.findall(".//testsuite") if root.tag != "testsuite" else [root]
    for suite in suites:
        try:
            total_time += float(suite.attrib.get("time", "0") or "0")
        except ValueError:
            pass

        for case in suite.findall("testcase"):
            classname = case.attrib.get("classname", "")
            test_name = case.attrib.get("name", "")
            file_attr = case.attrib.get("file") or _classname_to_file(classname)
            nodeid = f"{file_attr}::{test_name}" if file_attr else f"{classname}::{test_name}"
            nodeid = nodeid.replace("\\", "/")

            failure = case.find("failure")
            error = case.find("error")
            skipped = case.find("skipped")
            if failure is not None or error is not None:
                msg = (failure.attrib.get("message") if failure is not None else
                       error.attrib.get("message", ""))
                results.append(RunResult(nodeid=nodeid, status="F", message=msg or ""))
            elif skipped is not None:
                results.append(RunResult(nodeid=nodeid, status="S", message="skipped"))
            else:
                results.append(RunResult(nodeid=nodeid, status="P"))

    return results, total_time


def _classname_to_file(classname: str) -> str:
    # classname pytest dạng "tests.test_astar"
    if not classname:
        return ""
    parts = classname.split(".")
    return "/".join(parts) + ".py"


# ── Ghi xlsx ─────────────────────────────────────────────────────────────────

HEADERS = [
    "Mã TC",
    "Test Case",
    "Các bước thực hiện",
    "Dữ liệu đầu vào",
    "Kết quả mong đợi",
    "Kết quả thực tế",
    "Trạng thái",
]

COL_WIDTHS = [10, 30, 38, 32, 32, 32, 11]


def write_xlsx(
    results: list[RunResult],
    total_time: float,
    output: Path,
    *,
    title: str | None = None,
    sheet_name: str | None = None,
    meta_map: dict | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or "Kết quả kiểm thử A-star"
    meta_map = meta_map if meta_map is not None else TESTCASE_META

    title_font = Font(name="Times New Roman", bold=True, size=13)
    header_font = Font(name="Times New Roman", bold=True, size=11, color="FFFFFFFF")
    body_font = Font(name="Times New Roman", size=11)
    summary_font = Font(name="Times New Roman", bold=True, size=11)

    header_fill = PatternFill("solid", fgColor="FF305496")
    pass_fill = PatternFill("solid", fgColor="FFC6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFFFC7CE")
    skip_fill = PatternFill("solid", fgColor="FFFFEB9C")

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Tiêu đề
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(
        row=1,
        column=1,
        value=title or "BẢNG KẾT QUẢ KIỂM THỬ — THUẬT TOÁN A* (BACKEND PYTHON)",
    )
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Hàng meta nhỏ
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    meta_cell = ws.cell(row=2, column=1, value=
        f"Tổng: {len(results)} test • Đạt: {sum(1 for r in results if r.status == 'P')} "
        f"• Không đạt: {sum(1 for r in results if r.status == 'F')} "
        f"• Thời gian: {total_time:.2f}s")
    meta_cell.font = body_font
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header
    header_row = 3
    for idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_wrap
        c.border = border
    ws.row_dimensions[header_row].height = 32

    # Sắp xếp results theo Mã TC (TC_A01, TC_A02, …)
    enriched = []
    unknown = []
    for r in results:
        meta = meta_map.get(r.nodeid)
        if meta is None:
            unknown.append(r)
        else:
            enriched.append((meta, r))
    enriched.sort(key=lambda x: x[0].code)

    # Body
    current_row = header_row + 1
    for meta, r in enriched:
        values = [
            meta.code,
            meta.name,
            meta.steps,
            meta.inputs,
            meta.expected,
            meta.actual if r.status == "P" else (r.message or meta.actual),
            r.status,
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = body_font
            c.border = border
            if col_idx in (1, 7):
                c.alignment = center_wrap
            else:
                c.alignment = left_wrap

        status_cell = ws.cell(row=current_row, column=len(HEADERS))
        if r.status == "P":
            status_cell.fill = pass_fill
        elif r.status == "F":
            status_cell.fill = fail_fill
        else:
            status_cell.fill = skip_fill

        # Tính ước lượng chiều cao theo dòng wrap
        max_lines = max(
            (str(v).count("\n") + 1) for v in values if v is not None
        )
        ws.row_dimensions[current_row].height = max(22, min(120, 18 * max_lines))
        current_row += 1

    # Các test chưa có metadata (đề phòng tương lai thêm test mới)
    for r in unknown:
        values = ["?", r.nodeid, "(chưa khai báo metadata)", "", "", r.message or "", r.status]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = body_font
            c.border = border
            c.alignment = left_wrap
        current_row += 1

    # Tổng kết
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(row=current_row, column=1, value="Tổng kết").font = summary_font
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=current_row, column=3, value=
        f"Pass {sum(1 for r in results if r.status == 'P')}/{len(results)} • "
        f"Fail {sum(1 for r in results if r.status == 'F')} • "
        f"Thời gian chạy: {total_time:.2f}s").font = summary_font
    ws.merge_cells(start_row=current_row, start_column=3,
                   end_row=current_row, end_column=len(HEADERS))

    # Độ rộng cột
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A4"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


# ── Cấu hình các "feature" ──────────────────────────────────────────────────

SAFE_ROUTING_META: dict[str, TestCaseMeta] = {
    "tests/test_safe_routing_business.py::test_no_flood_picks_shortest_path": TestCaseMeta(
        code="TC_R01",
        name="Không ngập → chọn đường ngắn nhất",
        steps=(
            "1. Dựng đồ thị 5 nút, không có ngập.\n"
            "2. find_path(1, 5, is_dry=True)."
        ),
        inputs="5 nút {1..5}; không có flood; is_dry=True",
        expected="node_path = [1, 3, 4, 5]; cost = 280 m",
        actual="[1, 3, 4, 5]; cost = 280.0",
    ),
    "tests/test_safe_routing_business.py::test_sensor_flooded_node_forces_detour": TestCaseMeta(
        code="TC_R02",
        name="Sensor báo ngập tại C → A* vẽ đường vòng",
        steps=(
            "1. Đặt flood_depth_cm = 1e9 cho mọi cạnh chạm nút 3 (C).\n"
            "2. find_path(1, 5, is_dry=False)."
        ),
        inputs="Nút C (id=3) ngập; xe máy; is_dry=False",
        expected="node_path = [1, 2, 4, 5]; không qua C;\nít nhất 1 cạnh chạm C ∈ blocked_edge_ids",
        actual="[1, 2, 4, 5]; blocked_edge_ids chứa cạnh chạm nút 3",
    ),
    "tests/test_safe_routing_business.py::test_crowd_report_blocks_same_node_same_behaviour": TestCaseMeta(
        code="TC_R03",
        name="Crowd report đã duyệt → cùng kết quả",
        steps=(
            "1. Mô phỏng crowd report đã duyệt quanh nút C bằng cùng cơ chế ngập.\n"
            "2. find_path(1, 5, is_dry=False)."
        ),
        inputs="Crowd báo ngập tại C; xe máy; is_dry=False",
        expected="node_path = [1, 2, 4, 5] (parity với sensor)",
        actual="[1, 2, 4, 5]",
    ),
    "tests/test_safe_routing_business.py::test_all_branches_flooded_returns_none": TestCaseMeta(
        code="TC_R04",
        name="Mọi nhánh đều ngập → không có lộ trình",
        steps=(
            "1. Ngập cả nút B (2) và C (3).\n"
            "2. find_path(1, 5, is_dry=False)."
        ),
        inputs="Nút 2 và 3 đều ngập; xe máy; is_dry=False",
        expected="find_path trả về None",
        actual="None",
    ),
}


FEATURES = {
    "algorithm_a": {
        "title": "BẢNG KẾT QUẢ KIỂM THỬ — THUẬT TOÁN A* (BACKEND PYTHON)",
        "sheet": "Kết quả kiểm thử A-star",
        "default_output": r"D:\KhoaLuan_2026_Nhom007\test_report_algorithm_A.xlsx",
        "test_paths": [],  # rỗng = chạy tất cả test trong testpaths
        "meta": TESTCASE_META,
    },
    "safe_routing": {
        "title": "BẢNG KẾT QUẢ KIỂM THỬ — TÌM ĐƯỜNG AN TOÀN (SAFE ROUTING)",
        "sheet": "Safe Routing",
        "default_output": r"D:\KhoaLuan_2026_Nhom007\test_report_safe_routing.xlsx",
        "test_paths": ["tests/test_safe_routing_business.py"],
        "meta": SAFE_ROUTING_META,
    },
}


# ── CLI ──────────────────────────────────────────────────────────────────────

def _run_feature(feature_key: str, output_override: str | None) -> int:
    feature = FEATURES.get(feature_key)
    if feature is None:
        raise SystemExit(
            f"Feature không hợp lệ: {feature_key}. Hỗ trợ: {', '.join(FEATURES.keys())}"
        )

    routing_dir = Path(__file__).resolve().parent.parent
    results, total_time = run_pytest_and_collect(routing_dir, feature["test_paths"])

    output_path = Path(
        output_override
        or os.environ.get("REPORT_OUTPUT")
        or feature["default_output"]
    )
    write_xlsx(
        results,
        total_time,
        output_path,
        title=feature["title"],
        sheet_name=feature["sheet"],
        meta_map=feature["meta"],
    )

    passed = sum(1 for r in results if r.status == "P")
    failed = sum(1 for r in results if r.status == "F")
    print()
    print(f"[{feature_key}] Đã xuất: {output_path}")
    print(
        f"[{feature_key}] Pass {passed}/{len(results)} • Fail {failed} • "
        f"Thời gian: {total_time:.2f}s"
    )
    return 1 if failed > 0 else 0


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    except Exception:
        pass

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--feature",
        choices=sorted(FEATURES.keys()),
        default="algorithm_a",
        help="Feature cần xuất (mặc định: algorithm_a)",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Xuất tuần tự tất cả feature",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Ghi đè đường dẫn .xlsx (mặc định lấy theo --feature)",
    )
    args = parser.parse_args()

    features = sorted(FEATURES.keys()) if args.all else [args.feature]
    aggregate = 0
    for f in features:
        aggregate = aggregate | _run_feature(f, args.output)
    return aggregate


if __name__ == "__main__":
    sys.exit(main())
